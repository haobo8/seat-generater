from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment, PatternFill
import datetime
stag=open("seat.txt")
stagstr=stag.read()
seat = Workbook()
ws = seat.active
ws.title="座位表"
ws.merge_cells('A1:M1')
title_font = Font(name='宋体', size=24)
timenext= datetime.date.today()+datetime.timedelta(weeks=2)
weekn=timenext.strftime("%Y年 %W周")
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws['A1'].font = title_font
ws['A1']="ICC S1C5座位表 "+weekn+" 班主任：王思思"
dest_filename = 'ICC S1C5 座位表.xlsx'
ws.merge_cells('D2:J2')
subtitle_font = Font(name='宋体', size=16)
ws['D2'].alignment = Alignment(horizontal='center', vertical='center')
ws['D2'].font = subtitle_font
ws['D2']='白板'
yelow = PatternFill("solid", fgColor="FFFF00")
ws.merge_cells('L2:L3')
ws['L2'].alignment = Alignment(horizontal='center', vertical='center')
ws['L2'].font = subtitle_font
ws['L2']='门'
ws['L2'].fill=yelow
subsubtitle_font = Font(name='宋体', size=11)
ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
ws['C3'].font = subsubtitle_font
ws['C3']='讲台'
ws['C3'].fill=yelow
content_font = Font(name='宋体', size=18)
nrow=4
ncol=2
nstr=""
for i in stagstr:
    if i=="\t":
        ws.cell(nrow,ncol).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(nrow,ncol).font = content_font
        ws.cell(nrow,ncol).value=nstr
        nstr=""
        ncol+=1
    if i=="\n":
        nstr=""
        ncol=2
        nrow+=1
    if i!="\n" and i!="\t":
        nstr+=i
ws.merge_cells('D4:D7')
small_font = Font(name='宋体', size=6)
ws['D4'].alignment = Alignment(horizontal='center', vertical='center')
ws['D4'].font = small_font
ws['D4']='过道'
ws.merge_cells('G4:G6')
ws['G4'].alignment = Alignment(horizontal='center', vertical='center')
ws['G4'].font = small_font
ws['G4']='过道'
ws.merge_cells('J4:J7')
ws['J4'].alignment = Alignment(horizontal='center', vertical='center')
ws['J4'].font = small_font
ws['J4']='过道'
for row in range(4,8):
    ws.cell(row,1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row,1).font = content_font
    ws.cell(row,1).value='窗户'
    ws.cell(row,1).fill=yelow
    ws.cell(row,13).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row,13).font = content_font
    ws.cell(row,13).value='墙'
    ws.cell(row,13).fill=yelow
ws['L4'].fill=yelow
seat.save(filename = dest_filename)